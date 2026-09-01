"""
Lemon Clean 清潔承攬 — 場次時數 / 承攬費申報 / 工具包押金 / 介紹獎金 / 元大帳戶
檔案：modules/cleaning_process_4.py

公開函式：
    run_session_hours()   場次時數（cleaning_file_id 寫入 salary_id「場次和時數」）
    run_contract_report() 承攬費申報（寫入 contract_report_id／contract_report_sheet）
    run_tool_deposit()    工具包押金＋介紹獎金
    run_yuanta()          元大帳戶 / 元大工具包押金 xlsx

三者為各自獨立的作業，彼此不再互相呼叫；執行「工具包押金」不會
連帶執行「場次時數」或「承攬費申報」。

工作日規則：
    10 日 / 20 日若遇週六、週日或台灣國定假日，
    一律往前推至最近工作日。
"""

from __future__ import annotations

import datetime
import io
import time
from typing import List

import gspread
import holidays

from modules.auth import get_gspread_client
from modules.master_sheet import record_execution
from modules.period_utils import format_taipei_time


TS_FMT = "%Y/%m/%d %H:%M"
DEPOSIT_THRESHOLD = 80
DEPOSIT_BY_REGION = {
    "台北": 2000,
    "桃園": 2000,
    "新竹": 2000,
    "台中": 1500,
}
DEPOSIT_OTHER = 2000
INTRO_BONUS = 1000
TOOL_DEPOSIT_START_ROW = 121


OPEN_RETRY_DELAYS = (5, 10)  # 秒；PermissionError 重試等待時間


def _open_by_key_with_retry(gc: gspread.Client, file_id: str) -> gspread.Spreadsheet:
    """
    開啟試算表；遇到 PermissionError 時重試。

    gspread 對 Google Sheets API 回傳的 403 一律轉成內建 PermissionError，
    但 403 不只代表真的沒權限，短時間內請求過多觸發的額度限制
    （rate limit / quota exceeded）也會回傳 403，屬於暫時性錯誤，
    等幾秒重試通常就會成功，不代表分享設定有問題。
    """
    last_err: PermissionError | None = None
    for attempt, delay in enumerate((0,) + OPEN_RETRY_DELAYS):
        if delay:
            time.sleep(delay)
        try:
            return gc.open_by_key(file_id)
        except PermissionError as e:
            last_err = e
    raise last_err


def _now_ts() -> str:
    return format_taipei_time(fmt=TS_FMT)


def _log(log: List[str], msg: str) -> None:
    log.append(msg)


def _to_num(val) -> float:
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _col_letter(n: int) -> str:
    result = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def _previous_taiwan_business_day(target: datetime.date) -> datetime.date:
    """週末或台灣國定假日，往前找到最近工作日。"""
    tw = holidays.country_holidays("TW", years=[target.year])
    d = target
    while d.weekday() >= 5 or d in tw:
        d -= datetime.timedelta(days=1)
    return d


def _next_month_tenth(period: str) -> datetime.date:
    year = int(period[:4])
    month = int(period[4:6])
    if month == 12:
        target = datetime.date(year + 1, 1, 10)
    else:
        target = datetime.date(year, month + 1, 10)
    return _previous_taiwan_business_day(target)


def _yuanta_target_date(period: str, is_first_half: bool) -> datetime.date:
    """
    -1：當月 20 日；-2：隔月 10 日。
    週末或台灣國定假日均提前。
    """
    year = int(period[:4])
    month = int(period[4:6])
    if is_first_half:
        target = datetime.date(year, month, 20)
    else:
        if month == 12:
            target = datetime.date(year + 1, 1, 10)
        else:
            target = datetime.date(year, month + 1, 10)
    return _previous_taiwan_business_day(target)


# ============================================================
# 場次時數
# ============================================================

def run_session_hours(
    cleaning_file_id: str,
    region: str,
    period: str,
    is_first_half: bool,
    log: List[str],
    region_cfg: dict = None,
    **kwargs,
) -> bool:
    """將清潔承攬檔案 ID 寫入 salary_id 試算表的「場次和時數」分頁。"""
    label = "上半月" if is_first_half else "下半月"
    _log(log, f"▶ 場次時數 {label} 開始")

    if is_first_half:
        _log(log, "  上半月不需寫入場次時數，略過")
        _log(log, f"✅ 場次時數 {label} 完成｜{_now_ts()}")
        return True

    try:
        salary_id = str((region_cfg or {}).get("salary_id", "") or "").strip()
        if not salary_id:
            raise ValueError("地區設定缺少 salary_id")

        gc = get_gspread_client()
        salary_ss = _open_by_key_with_retry(gc, salary_id)
        ws_counts = salary_ss.worksheet("場次和時數")

        month = int(period[4:6])
        id_col = 5 + (month - 1) * 3
        ws_counts.update_cell(1, id_col, cleaning_file_id)
        _log(log, f"  清潔承攬 ID 已寫入 場次和時數!{_col_letter(id_col)}1")

        record_execution(region, period, "場次時數", None)
        _log(log, f"✅ 場次時數 {label} 完成｜{_now_ts()}")
        return True

    except Exception as e:
        detail = str(e).strip() or repr(e)
        _log(log, f"❌ 場次時數失敗：{detail}")
        return False


# ============================================================
# 承攬費申報
# ============================================================

def run_contract_report(
    cleaning_file_id: str,
    region: str,
    period: str,
    is_first_half: bool,
    log: List[str],
    region_cfg: dict = None,
    **kwargs,
) -> bool:
    """
    將清潔承攬檔案 ID 寫入地區設定 contract_report_id／contract_report_sheet
    指定的試算表分頁：B{月份+1} = 該月下半月（-2）清潔承攬檔案 ID。
    例：B2=YYYY01-2、B3=YYYY02-2、B4=YYYY03-2 ……以此類推。
    """
    label = "上半月" if is_first_half else "下半月"
    _log(log, f"▶ 承攬費申報 {label} 開始")

    if is_first_half:
        _log(log, "  上半月不需寫入承攬費申報，略過")
        _log(log, f"✅ 承攬費申報 {label} 完成｜{_now_ts()}")
        return True

    cfg = region_cfg or {}
    contract_report_id = str(cfg.get("contract_report_id", "") or "").strip()
    contract_report_sheet = str(cfg.get("contract_report_sheet", "") or "").strip()

    missing = []
    if not contract_report_id:
        missing.append("contract_report_id")
    if not contract_report_sheet:
        missing.append("contract_report_sheet")

    if missing:
        _log(
            log,
            f"⚠️ 「{region}」地區設定缺少 {'、'.join(missing)}，"
            "請至「地區設定」回填 contract_report_id／contract_report_sheet 後再執行承攬費申報",
        )
        return False

    try:
        gc = get_gspread_client()
        report_ss = _open_by_key_with_retry(gc, contract_report_id)
        ws_report = report_ss.worksheet(contract_report_sheet)

        month = int(period[4:6])
        row = month + 1
        ws_report.update_cell(row, 2, cleaning_file_id)
        _log(log, f"  清潔承攬 ID 已寫入 {contract_report_sheet}!B{row}")

        record_execution(region, period, "承攬費申報", None)
        _log(log, f"✅ 承攬費申報 {label} 完成｜{_now_ts()}")
        return True

    except Exception as e:
        detail = str(e).strip() or repr(e)
        _log(log, f"❌ 承攬費申報失敗：{detail}")
        return False


# ============================================================
# 工具包押金 / 介紹獎金
# ============================================================

def run_tool_deposit(
    cleaning_file_id: str,
    region: str,
    period: str,
    is_first_half: bool,
    log: List[str],
    region_cfg: dict = None,
    **kwargs,
) -> bool:
    label = "上半月" if is_first_half else "下半月"
    _log(log, f"▶ 工具包押金 & 介紹獎金 {label} 開始")

    try:
        gc = get_gspread_client()
        cleaning_ss = _open_by_key_with_retry(gc, cleaning_file_id)
        ws_summary = cleaning_ss.worksheet("場次時數薪資總表")
        ws_intro = cleaning_ss.worksheet("介紹獎金")

        if is_first_half:
            _tool_clear(ws_summary, ws_intro, log)
            dep_count = 0
        else:
            salary_id = str((region_cfg or {}).get("salary_id", "") or "").strip()
            if not salary_id:
                raise ValueError("地區設定缺少 salary_id")

            salary_ss = _open_by_key_with_retry(gc, salary_id)
            ws_deposit = salary_ss.worksheet("工具包押金")

            amount = next(
                (v for k, v in DEPOSIT_BY_REGION.items() if k in region),
                DEPOSIT_OTHER,
            )
            dep_count = _tool_process_v2(
                ws_deposit=ws_deposit,
                ws_summary=ws_summary,
                ws_intro=ws_intro,
                deposit_amount=amount,
                period=period,
                log=log,
            )

        record_execution(region, period, "工具包押金", dep_count)
        _log(log, f"✅ 工具包押金 {label} 完成｜{_now_ts()}")
        return True

    except Exception as e:
        detail = str(e).strip() or repr(e)
        _log(log, f"❌ 工具包押金失敗：{detail}")
        return False


def _tool_clear(
    ws_summary: gspread.Worksheet,
    ws_intro: gspread.Worksheet,
    log: List[str],
) -> None:
    ws_summary.batch_clear([
        f"A{TOOL_DEPOSIT_START_ROW}:E",
        "AB4:AE",
    ])
    ws_intro.batch_clear(["A2:C"])
    _log(log, "  上半月：已清空 A121:E、AB4:AE、介紹獎金 A2:C")


def _tool_process_v2(
    ws_deposit: gspread.Worksheet,
    ws_summary: gspread.Worksheet,
    ws_intro: gspread.Worksheet,
    deposit_amount: int,
    period: str,
    log: List[str],
) -> int:
    rows = ws_deposit.get("A2:J") or []
    due_text = _next_month_tenth(period).strftime("%Y/%m/%d")

    selected = []
    due_updates = []

    for sheet_row, row in enumerate(rows, start=2):
        row = list(row) + [""] * (10 - len(row))
        name = str(row[0]).strip()
        i_value = _to_num(row[8])
        current_due = str(row[6]).strip().replace("-", "/")

        if current_due:
            try:
                current_due = datetime.datetime.strptime(
                    current_due, "%Y/%m/%d"
                ).strftime("%Y/%m/%d")
            except ValueError:
                pass

        if name and i_value >= DEPOSIT_THRESHOLD and (
            not current_due or current_due == due_text
        ):
            i_out = int(i_value) if float(i_value).is_integer() else i_value
            selected.append((name, i_out))
            if not current_due:
                due_updates.append({
                    "range": f"'{ws_deposit.title}'!G{sheet_row}",
                    "values": [[due_text]],
                })

    if due_updates:
        ws_deposit.spreadsheet.values_batch_update({
            "valueInputOption": "USER_ENTERED",
            "data": due_updates,
        })

    # 介紹獎金：符合工具包資格且 J 欄有介紹人
    selected_names = {name for name, _ in selected}
    intro_rows = []
    for row in rows:
        row = list(row) + [""] * (10 - len(row))
        name = str(row[0]).strip()
        introducer = str(row[9]).strip()
        if name in selected_names and introducer:
            intro_rows.append([name, introducer, INTRO_BONUS])

    ws_intro.batch_clear(["A2:C"])
    if intro_rows:
        ws_intro.update(
            f"A2:C{1 + len(intro_rows)}",
            intro_rows,
            value_input_option="USER_ENTERED",
        )
    _log(log, f"  介紹獎金回填 {len(intro_rows)} 筆")

    # 清掉上一期工具包資料，但不動 A1:B120
    ws_summary.batch_clear(["A121:E", "AB4:AE120"])

    if selected:
        end_row = TOOL_DEPOSIT_START_ROW + len(selected) - 1
        ws_summary.update(
            f"A{TOOL_DEPOSIT_START_ROW}:B{end_row}",
            [[name, count] for name, count in selected],
            value_input_option="USER_ENTERED",
        )
        _log(log, f"  工具包押金名單寫入 A121:B{end_row}，共 {len(selected)} 筆")

        # AD=押金金額，AE=姓名
        end = 3 + len(selected)
        ws_summary.update(
            f"AD4:AE{end}",
            [[deposit_amount, name] for name, _ in selected],
            value_input_option="USER_ENTERED",
        )

        # AB/AC 依 AE 姓名比對 H 欄並帶入 I/J
        hij = ws_summary.get("H4:J120") or []
        account_map = {}
        for r in hij:
            if not r:
                continue
            h_name = str(r[0]).strip() if len(r) > 0 else ""
            if h_name:
                account_map[h_name] = (
                    r[1] if len(r) > 1 else "",
                    r[2] if len(r) > 2 else "",
                )

        ae_values = ws_summary.get(f"AE4:AE{end}") or []
        ab_ac = []
        missing = []
        for i in range(len(selected)):
            ae_name = (
                str(ae_values[i][0]).strip()
                if i < len(ae_values) and ae_values[i]
                else ""
            )
            i_val, j_val = account_map.get(ae_name, ("", ""))
            ab_ac.append([i_val, j_val])
            if ae_name and ae_name not in account_map:
                missing.append(ae_name)

        ws_summary.update(
            f"AB4:AC{end}",
            ab_ac,
            value_input_option="USER_ENTERED",
        )
        if missing:
            _log(log, "  ⚠️ H欄找不到姓名：" + "、".join(missing))

    _log(log, f"  工具包押金完成：{len(selected)} 筆；提領日 {due_text}")
    return len(selected)


# ============================================================
# 元大帳戶
# ============================================================

def _yuanta_find_period_file(
    root_folder_id: str,
    period: str,
    file_name: str,
) -> tuple[str, str]:
    from modules.auth import get_drive_service

    drive = get_drive_service()

    def query(parent_id: str, name: str, mime_type: str):
        resp = drive.files().list(
            q=(
                f"'{parent_id}' in parents and name = '{name}' "
                f"and mimeType = '{mime_type}' and trashed = false"
            ),
            fields="files(id,name)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
            pageSize=10,
        ).execute()
        files = resp.get("files", [])
        return files[0]["id"] if files else None

    period_folder_id = query(
        root_folder_id,
        period,
        "application/vnd.google-apps.folder",
    )
    if not period_folder_id:
        raise FileNotFoundError(f"找不到期別資料夾：{period}")

    file_id = query(
        period_folder_id,
        file_name,
        "application/vnd.google-apps.spreadsheet",
    )
    if not file_id:
        raise FileNotFoundError(f"找不到試算表：{file_name}")

    return period_folder_id, file_id


def _yuanta_find_other_file(
    root_folder_id: str,
    period: str,
    region: str,
) -> tuple[str, str]:
    return _yuanta_find_period_file(
        root_folder_id,
        period,
        f"{period}其他承攬-{region}",
    )


def _yuanta_nonempty_rows(rows: list[list], width: int = 4) -> list[list]:
    result = []
    for row in rows:
        padded = list(row[:width]) + [""] * max(0, width - len(row))
        if any(str(v).strip() for v in padded):
            result.append(padded[:width])
    return result


def _yuanta_wait_values(
    ws: gspread.Worksheet,
    a1_range: str,
    expected_rows: list[list],
    log: List[str],
    timeout: int = 30,
) -> None:
    deadline = time.time() + timeout
    expected_count = len(expected_rows)

    while time.time() < deadline:
        actual = ws.get(a1_range, value_render_option="UNFORMATTED_VALUE") or []
        actual_count = sum(
            1 for row in actual if any(str(v).strip() for v in row)
        )
        if actual_count >= expected_count:
            _log(log, f"  已確認 {ws.title}!{a1_range} 寫入完成（{actual_count} 筆）")
            time.sleep(2)
            return
        time.sleep(2)

    raise TimeoutError(f"{ws.title}!{a1_range} 寫入後仍未同步，取消匯出")


def _yuanta_export_xlsx(
    spreadsheet_id: str,
    folder_id: str,
    output_name: str,
    log: List[str],
) -> None:
    from googleapiclient.http import MediaIoBaseUpload
    from openpyxl import load_workbook
    from modules.auth import get_drive_service, get_jenny_drive_service

    source_drive = get_drive_service()
    drive = get_jenny_drive_service()

    request = source_drive.files().export_media(
        fileId=spreadsheet_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    data = request.execute()
    _log(log, f"  xlsx 下載完成：{len(data)} bytes")

    # 只後處理匯出的 xlsx，不修改 Google Sheet
    wb = load_workbook(io.BytesIO(data))
    if "元大" not in wb.sheetnames:
        raise RuntimeError("匯出的 xlsx 找不到『元大』工作表")

    ws = wb["元大"]
    last_row = max(ws.max_row, 3)
    for row_idx in range(3, last_row + 1):
        for col_idx in range(6, 13):  # F:L
            ws.cell(row=row_idx, column=col_idx).value = None
        ws.cell(row=row_idx, column=4).number_format = "#,##0"

    stream = io.BytesIO()
    wb.save(stream)
    data = stream.getvalue()

    safe_name = output_name.replace("'", "\\'")
    q = f"'{folder_id}' in parents and name = '{safe_name}' and trashed = false"
    existing = drive.files().list(
        q=q,
        fields="files(id,name,modifiedTime)",
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
        orderBy="modifiedTime desc",
        pageSize=100,
    ).execute().get("files", [])

    media = MediaIoBaseUpload(
        io.BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )

    if existing:
        drive.files().update(
            fileId=existing[0]["id"],
            media_body=media,
            body={"name": output_name},
            supportsAllDrives=True,
            fields="id,name",
        ).execute()
        _log(log, f"  xlsx 已覆蓋：{output_name}")
    else:
        drive.files().create(
            body={"name": output_name, "parents": [folder_id]},
            media_body=media,
            fields="id,name",
            supportsAllDrives=True,
        ).execute()
        _log(log, f"  xlsx 已建立：{output_name}")


def run_yuanta(
    cleaning_file_id: str,
    region: str,
    period: str,
    is_first_half: bool,
    log: List[str],
    region_cfg: dict = None,
    **kwargs,
) -> bool:
    if period.endswith("-1"):
        is_first_half = True
    elif period.endswith("-2"):
        is_first_half = False
    else:
        _log(log, f"❌ 元大帳戶失敗：期別格式錯誤 {period}")
        return False

    label = "上半月" if is_first_half else "下半月"
    _log(log, f"▶ 元大帳戶 {label} 開始（{period}）")

    try:
        cfg = region_cfg or {}
        root_folder_id = str(cfg.get("root_folder_id", "") or "").strip()
        if not root_folder_id:
            raise ValueError("地區設定缺少 root_folder_id")

        gc = get_gspread_client()
        cleaning_ss = gc.open_by_key(cleaning_file_id)
        ws_summary = cleaning_ss.worksheet("場次時數薪資總表")

        yuanta_name = f"{period}元大帳戶-{region}"
        period_folder_id, yuanta_file_id = _yuanta_find_period_file(
            root_folder_id,
            period,
            yuanta_name,
        )
        yuanta_ss = gc.open_by_key(yuanta_file_id)
        ws_all = yuanta_ss.worksheet("all")
        ws_yuanta = yuanta_ss.worksheet("元大")

        source_range = "N4:Q" if is_first_half else "U4:X"
        other_source_range = "N3:Q" if is_first_half else "U3:X"

        cleaning_rows = _yuanta_nonempty_rows(
            ws_summary.get(source_range, value_render_option="UNFORMATTED_VALUE") or []
        )

        ws_all.batch_clear(["A2:D"])
        next_row = 2
        if cleaning_rows:
            end = next_row + len(cleaning_rows) - 1
            ws_all.update(
                f"A{next_row}:D{end}",
                cleaning_rows,
                value_input_option="USER_ENTERED",
            )
            next_row = end + 1

        _, other_file_id = _yuanta_find_other_file(
            root_folder_id,
            period,
            region,
        )
        other_ss = gc.open_by_key(other_file_id)
        other_ws = other_ss.worksheet("薪資總表")
        other_rows = _yuanta_nonempty_rows(
            other_ws.get(
                other_source_range,
                value_render_option="UNFORMATTED_VALUE",
            ) or []
        )

        if other_rows:
            other_end = next_row + len(other_rows) - 1
            ws_all.update(
                f"A{next_row}:D{other_end}",
                other_rows,
                value_input_option="USER_ENTERED",
            )

        all_rows = cleaning_rows + other_rows
        if all_rows:
            _yuanta_wait_values(
                ws_all,
                f"A2:D{1 + len(all_rows)}",
                all_rows,
                log,
            )

        # 排除 B 空白與現金
        bank_rows = [
            row for row in all_rows
            if str(row[1] if len(row) > 1 else "").strip()
            and str(row[1]).strip() != "現金"
        ]

        ws_yuanta.batch_clear(["A3:E"])
        yyyymm = period[:6]
        ws_yuanta.update("H2", [[yyyymm]], value_input_option="USER_ENTERED")

        if bank_rows:
            target_date = _yuanta_target_date(period, is_first_half)
            end = 2 + len(bank_rows)
            export_rows = [
                [target_date.strftime("%Y%m%d")] + list(row[:4])
                for row in bank_rows
            ]
            ws_yuanta.update(
                f"A3:E{end}",
                export_rows,
                value_input_option="USER_ENTERED",
            )
            _yuanta_wait_values(ws_yuanta, f"A3:E{end}", export_rows, log)
            _log(log, f"  承攬費 {len(bank_rows)} 筆；入帳日 {target_date:%Y%m%d}")

        fee_name = f"{period}元大承攬費-{region}.xlsx"
        _yuanta_export_xlsx(
            yuanta_file_id,
            period_folder_id,
            fee_name,
            log,
        )

        # 下半月若 A121 非空，另外輸出工具包押金
        if not is_first_half:
            a121 = str(ws_summary.acell("A121").value or "").strip()
            if a121:
                deposit_rows = _yuanta_nonempty_rows(
                    ws_summary.get(
                        "AB4:AE",
                        value_render_option="UNFORMATTED_VALUE",
                    ) or []
                )
                if deposit_rows:
                    ws_yuanta.batch_clear(["A3:E"])
                    deposit_date = _next_month_tenth(period)
                    end = 2 + len(deposit_rows)
                    deposit_export_rows = [
                        [deposit_date.strftime("%Y%m%d")] + list(row[:4])
                        for row in deposit_rows
                    ]
                    ws_yuanta.update(
                        f"A3:E{end}",
                        deposit_export_rows,
                        value_input_option="USER_ENTERED",
                    )
                    _yuanta_wait_values(
                        ws_yuanta,
                        f"A3:E{end}",
                        deposit_export_rows,
                        log,
                    )
                    deposit_name = f"{period}元大工具包押金-{region}.xlsx"
                    _yuanta_export_xlsx(
                        yuanta_file_id,
                        period_folder_id,
                        deposit_name,
                        log,
                    )
                    _log(
                        log,
                        f"  工具包押金 {len(deposit_rows)} 筆；入帳日 {deposit_date:%Y%m%d}",
                    )

        record_execution(region, period, "元大帳戶", None)
        _log(log, f"✅ 元大帳戶 {label} 完成｜{_now_ts()}")
        return True

    except Exception as e:
        detail = str(e).strip() or repr(e)
        _log(log, f"❌ 元大帳戶失敗：{detail}")
        return False
